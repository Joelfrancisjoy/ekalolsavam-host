from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import PermissionDenied
from django.db.models import Q
from django.db import IntegrityError, transaction
from django.core.mail import send_mail
from django.core import signing
from django.utils import timezone
from rest_framework.parsers import MultiPartParser, FormParser
from drf_spectacular.utils import extend_schema, OpenApiResponse
from drf_spectacular.types import OpenApiTypes
import secrets
import re
import os
import csv
import io

from events.serializers import EventSerializer
from .models import User
from .workflow_models import (
    AdminIssuedID, SchoolParticipant, SchoolGroupEntry, SchoolGroupMember, SchoolVolunteerAssignment,
    SchoolStanding, IDSignupRequest
)
from .workflow_serializers import (
    AdminIssuedIDSerializer, SchoolParticipantSerializer,
    SchoolGroupEntrySerializer,
    SchoolVolunteerAssignmentSerializer, SchoolStandingSerializer,
    IDSignupRequestSerializer, StudentGroupProfileUpdateSerializer,
    StudentAllowedEventsResponseSerializer,
    SchoolIndividualEventsQuerySerializer,
    SchoolGroupEventsQuerySerializer,
    AdminSchoolGroupApproveRequestSerializer,
    AdminSchoolGroupRejectRequestSerializer,
    AdminSchoolGroupApproveResponseSerializer,
    AdminSchoolGroupRejectResponseSerializer,
)
from .permissions import IsAdminRole, IsAdminOrVolunteerRole, IsAdminOrStudentSignupVolunteer
from core.serializers import ErrorEnvelopeSerializer


def _error_payload(message, code=None, details=None):
    payload = {
        'error': str(message),
        'detail': str(message),
    }
    if code:
        payload['code'] = code
    if details is not None:
        payload['details'] = details
    return payload


def _error_response(message, http_status, code=None, details=None):
    return Response(
        _error_payload(message, code=code, details=details),
        status=http_status,
    )


def _ensure_school_user_linked_to_school_profile(school_user):
    if school_user is None:
        return None
    if getattr(school_user, 'school', None) is not None:
        return school_user.school

    try:
        from .models import School
        match = School.objects.filter(name__iexact=(school_user.username or '').strip()).first()
        if match is None:
            return None
        school_user.school = match
        school_user.save(update_fields=['school'])
        return match
    except Exception:
        return None


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_link_school_user_to_school_profile(request):
    username = (request.data.get('username') or '').strip()
    school_id = request.data.get('school_id')

    if not username:
        return Response({'error': 'username is required'}, status=status.HTTP_400_BAD_REQUEST)
    if not school_id:
        return Response({'error': 'school_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        school_id = int(school_id)
    except Exception:
        return Response({'error': 'school_id must be an integer'}, status=status.HTTP_400_BAD_REQUEST)

    from .models import School
    school_obj = School.objects.filter(pk=school_id).first()
    if school_obj is None:
        return Response({'error': 'School not found'}, status=status.HTTP_404_NOT_FOUND)

    school_user = User.objects.filter(username__iexact=username, role='school').first()
    if school_user is None:
        return Response({'error': 'School user not found'}, status=status.HTTP_404_NOT_FOUND)

    school_user.school = school_obj
    school_user.save(update_fields=['school'])

    return Response({
        'message': 'School user linked successfully',
        'user_id': school_user.id,
        'username': school_user.username,
        'school_id': school_obj.id,
        'school_name': school_obj.name,
    }, status=status.HTTP_200_OK)


@extend_schema(
    methods=['GET'],
    responses={
        200: StudentAllowedEventsResponseSerializer,
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_allowed_events(request):
    user = request.user
    if getattr(user, 'role', None) != 'student':
        return _error_response(
            'Only students can access this resource',
            status.HTTP_403_FORBIDDEN,
            code='forbidden',
        )

    participant = None
    registration_id = (getattr(user, 'registration_id', None) or '').strip()
    if registration_id.startswith('SPP-'):
        try:
            participant_pk = int(registration_id.split('-', 1)[1])
            participant = SchoolParticipant.objects.select_related('school').prefetch_related('events').filter(pk=participant_pk).first()
        except Exception:
            participant = None

    if participant is None and getattr(user, 'school', None) is not None:
        try:
            participant = SchoolParticipant.objects.select_related('school').prefetch_related('events').filter(
                school__school=user.school,
                first_name__iexact=(user.first_name or ''),
                last_name__iexact=(user.last_name or ''),
            ).order_by('-submitted_at').first()
        except Exception:
            participant = None

    event_ids = []
    participant_id = None
    if participant is not None:
        participant_id = participant.pk
        event_ids = list(participant.events.values_list('id', flat=True))

    group_entries_payload = []
    group_event_ids = set()
    full_name = _normalize_person_name(user.first_name, user.last_name)
    if getattr(user, 'school', None) is not None:
        group_entries_qs = SchoolGroupEntry.objects.prefetch_related('events').filter(
            school__school=user.school,
            status='approved',
        ).filter(
            Q(leader_user=user) |
            (Q(leader_user__isnull=True) & Q(leader_full_name__iexact=full_name))
        ).order_by('-submitted_at')

        for group_entry in group_entries_qs:
            group_ids = list(group_entry.events.values_list('id', flat=True))
            group_event_ids.update(group_ids)
            group_entries_payload.append({
                'group_entry_id': group_entry.id,
                'group_id': group_entry.group_id,
                'leader_full_name': group_entry.leader_full_name,
                'leader_user_id': group_entry.leader_user_id,
                'event_ids': group_ids,
                'status': group_entry.status,
            })

    return Response({
        'event_ids': event_ids,
        'participant_id': participant_id,
        'school_id': getattr(getattr(user, 'school', None), 'id', None),
        'group_event_ids': sorted(group_event_ids),
        'group_entries': group_entries_payload,
    })


def _student_managed_group_entries_queryset(user):
    if getattr(user, 'role', None) != 'student':
        return SchoolGroupEntry.objects.none()

    if getattr(user, 'school', None) is None:
        return SchoolGroupEntry.objects.none()

    full_name = _normalize_person_name(user.first_name, user.last_name)
    return SchoolGroupEntry.objects.select_related(
        'leader_user',
        'school',
    ).prefetch_related(
        'members',
        'events',
    ).filter(
        school__school=user.school,
        status='approved',
    ).filter(
        Q(leader_user=user) |
        (Q(leader_user__isnull=True) & Q(leader_full_name__iexact=full_name))
    ).order_by('-submitted_at')


def _school_managed_group_entries_queryset(user):
    if getattr(user, 'role', None) != 'school':
        return SchoolGroupEntry.objects.none()

    return SchoolGroupEntry.objects.select_related(
        'leader_user',
        'school',
    ).prefetch_related(
        'members',
        'events',
    ).filter(
        school=user,
    ).order_by('-submitted_at')


def _apply_group_members_payload_update(group_entry, participants_payload):
    member_qs = group_entry.members.all()
    members_by_id = {member.id: member for member in member_qs}
    members_by_order = {member.member_order: member for member in member_qs}
    touched_ids = set()

    for participant_payload in participants_payload:
        member = None
        member_id = participant_payload.get('id')
        member_order = participant_payload.get('member_order')
        if member_id not in [None, '']:
            member = members_by_id.get(member_id)
        if member is None and member_order not in [None, '']:
            member = members_by_order.get(member_order)
        if member is None:
            raise ValueError('Participant not found in this group')

        member.first_name = participant_payload['first_name']
        member.last_name = participant_payload['last_name']
        if 'gender' in participant_payload:
            member.gender = participant_payload['gender']
        if 'student_class' in participant_payload:
            member.student_class = participant_payload['student_class']
        if 'phone' in participant_payload:
            member.phone = participant_payload['phone']
        member.save(update_fields=['first_name', 'last_name', 'gender', 'student_class', 'phone'])
        touched_ids.add(member.id)

    if len(touched_ids) != group_entry.participant_count:
        raise ValueError('All participants in the group must be provided for update')

    leader_member = group_entry.members.filter(is_leader=True).order_by('member_order').first()
    if leader_member is None:
        raise ValueError('Leader details are missing for this group entry')

    return leader_member


def _apply_group_entry_profile_update(group_entry, validated_data, notes=None):
    update_fields = []
    if 'gender_category' in validated_data:
        group_entry.gender_category = validated_data['gender_category']
        update_fields.append('gender_category')
        update_fields.append('updated_at')

    participants_payload = validated_data.get('participants')
    if participants_payload is not None:
        leader_member = _apply_group_members_payload_update(group_entry, participants_payload)
        group_entry.leader_full_name = _normalize_person_name(leader_member.first_name, leader_member.last_name)
        if 'leader_full_name' not in update_fields:
            update_fields.append('leader_full_name')
        if 'updated_at' not in update_fields:
            update_fields.append('updated_at')

    if notes is not None:
        group_entry.review_notes = str(notes).strip()
        if 'review_notes' not in update_fields:
            update_fields.append('review_notes')
        if 'updated_at' not in update_fields:
            update_fields.append('updated_at')

    if update_fields:
        group_entry.save(update_fields=update_fields)


@extend_schema(
    methods=['GET'],
    responses={
        200: SchoolGroupEntrySerializer,
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@extend_schema(
    methods=['PATCH'],
    request=StudentGroupProfileUpdateSerializer,
    responses={
        200: SchoolGroupEntrySerializer,
        400: OpenApiResponse(response=OpenApiTypes.OBJECT),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def student_group_profile(request, group_entry_id):
    if getattr(request.user, 'role', None) != 'student':
        return Response(
            {'error': 'Only students can access this resource', 'detail': 'Only students can access this resource'},
            status=status.HTTP_403_FORBIDDEN,
        )

    group_entry = _student_managed_group_entries_queryset(request.user).filter(pk=group_entry_id).first()
    if group_entry is None:
        return Response(
            {'error': 'Group entry not found or not authorized', 'detail': 'Group entry not found or not authorized'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == 'GET':
        return Response(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data)

    serializer = StudentGroupProfileUpdateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    validated_data = serializer.validated_data

    try:
        with transaction.atomic():
            _apply_group_entry_profile_update(group_entry, validated_data, notes=validated_data.get('notes'))
    except ValueError as exc:
        message = str(exc)
        return Response({'error': message, 'detail': message}, status=status.HTTP_400_BAD_REQUEST)

    return Response(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data)


@extend_schema(
    request=StudentGroupProfileUpdateSerializer,
    responses={
        200: SchoolGroupEntrySerializer,
        400: OpenApiResponse(response=OpenApiTypes.OBJECT),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def school_update_group_participant(request, group_entry_id):
    if getattr(request.user, 'role', None) != 'school':
        return Response(
            {'error': 'Only school users can access this resource', 'detail': 'Only school users can access this resource'},
            status=status.HTTP_403_FORBIDDEN,
        )

    group_entry = _school_managed_group_entries_queryset(request.user).filter(pk=group_entry_id).first()
    if group_entry is None:
        return Response({'error': 'Group entry not found', 'detail': 'Group entry not found'}, status=status.HTTP_404_NOT_FOUND)

    if group_entry.status == 'approved':
        message = 'Approved group entries cannot be edited from the school dashboard'
        return Response(
            {'error': message, 'detail': message},
            status=status.HTTP_400_BAD_REQUEST,
        )

    serializer = StudentGroupProfileUpdateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    validated_data = serializer.validated_data

    try:
        with transaction.atomic():
            _apply_group_entry_profile_update(group_entry, validated_data, notes=validated_data.get('notes'))
    except ValueError as exc:
        message = str(exc)
        return Response({'error': message, 'detail': message}, status=status.HTTP_400_BAD_REQUEST)

    return Response(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data, status=status.HTTP_200_OK)


@extend_schema(
    request=StudentGroupProfileUpdateSerializer,
    responses={
        200: SchoolGroupEntrySerializer,
        400: OpenApiResponse(response=OpenApiTypes.OBJECT),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_update_school_group_participant(request, group_entry_id):
    try:
        group_entry = SchoolGroupEntry.objects.select_related('school', 'leader_user').prefetch_related('members', 'events').get(pk=group_entry_id)
    except SchoolGroupEntry.DoesNotExist:
        return Response({'error': 'Group entry not found', 'detail': 'Group entry not found'}, status=status.HTTP_404_NOT_FOUND)

    serializer = StudentGroupProfileUpdateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    validated_data = serializer.validated_data

    try:
        with transaction.atomic():
            notes = validated_data.get('notes')
            _apply_group_entry_profile_update(group_entry, validated_data, notes=notes)
    except ValueError as exc:
        message = str(exc)
        return Response({'error': message, 'detail': message}, status=status.HTTP_400_BAD_REQUEST)

    return Response(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data, status=status.HTTP_200_OK)


def _generate_unique_student_username(participant_id, school_id):
    base = (participant_id or '').strip().lower()
    base = re.sub(r'[^a-z0-9._-]+', '', base)
    if not base:
        base = 'student'

    candidates = [base]
    if school_id is not None:
        candidates.append(f"{base}-{school_id}")

    for candidate in candidates:
        if not User.objects.filter(username=candidate).exists():
            return candidate

    suffix = 2
    while suffix <= 50:
        if school_id is not None:
            candidate = f"{base}-{school_id}-{suffix}"
        else:
            candidate = f"{base}-{suffix}"
        if not User.objects.filter(username=candidate).exists():
            return candidate
        suffix += 1

    raise ValueError('Unable to generate unique username')


def _student_placeholder_email(username):
    domain = 'students.ekalolsavam.local'
    candidate = f"{username}@{domain}"
    if not User.objects.filter(email=candidate).exists():
        return candidate

    suffix = 2
    while suffix <= 50:
        candidate = f"{username}-{suffix}@{domain}"
        if not User.objects.filter(email=candidate).exists():
            return candidate
        suffix += 1

    token = secrets.token_hex(3)
    return f"{username}-{token}@{domain}"


def _school_participant_registration_id(participant_pk):
    return f"SPP-{participant_pk}"


def _school_group_leader_registration_id(group_entry_pk):
    return f"SGE-{group_entry_pk}"


def _normalize_student_gender(gender):
    normalized_gender = str(gender or '').strip().upper()
    if normalized_gender in {'BOYS', 'GIRLS'}:
        return normalized_gender
    return None


def _resolve_section_from_student_class(student_class):
    try:
        student_class_int = int(student_class)
    except (TypeError, ValueError):
        raise ValueError(f'Invalid student class: {student_class}')

    try:
        section = _student_class_to_level_code(student_class_int)
    except ValueError:
        raise ValueError(f'Invalid student class: {student_class}. Must be between 1-12.')

    return student_class_int, section


def _default_student_class_for_group_class(group_class):
    normalized_group_class = str(group_class or '').strip().upper()
    class_by_group_level = {
        'LP': 3,
        'UP': 5,
        'HS': 8,
        'HSS': 11,
    }
    if normalized_group_class not in class_by_group_level:
        raise ValueError('group_class must be one of LP, UP, HS, HSS')
    return class_by_group_level[normalized_group_class]


def _find_existing_school_student_user(participant_school, registration_ids=None, username_hint=None):
    registration_ids = registration_ids or []
    for registration_id in registration_ids:
        normalized_registration_id = str(registration_id or '').strip()
        if not normalized_registration_id:
            continue
        existing_user = User.objects.filter(
            role='student',
            school=participant_school,
            registration_id=normalized_registration_id,
        ).first()
        if existing_user is not None:
            return existing_user

    normalized_username_hint = str(username_hint or '').strip().lower()
    if normalized_username_hint:
        existing_user = User.objects.filter(
            role='student',
            school=participant_school,
            username=normalized_username_hint,
        ).first()
        if existing_user is not None:
            return existing_user

    return None


def _provision_school_student_account(
    school_user,
    first_name,
    last_name,
    student_class,
    gender,
    username_seed,
    primary_registration_id,
    legacy_registration_ids=None,
    existing_user_hint=None,
    reassign_registration_id_values=None,
):
    participant_school = getattr(school_user, 'school', None)
    if participant_school is None:
        participant_school = _ensure_school_user_linked_to_school_profile(school_user)
    if participant_school is None:
        raise ValueError('School account is not linked to a School profile')

    student_class_int, section = _resolve_section_from_student_class(student_class)
    normalized_gender = _normalize_student_gender(gender)

    existing_user = None
    if (
        existing_user_hint is not None
        and getattr(existing_user_hint, 'role', None) == 'student'
        and getattr(existing_user_hint, 'school_id', None) == participant_school.id
    ):
        existing_user = existing_user_hint

    if existing_user is None:
        candidate_registration_ids = [primary_registration_id]
        if legacy_registration_ids:
            candidate_registration_ids.extend(legacy_registration_ids)
        existing_user = _find_existing_school_student_user(
            participant_school=participant_school,
            registration_ids=candidate_registration_ids,
            username_hint=username_seed,
        )

    if reassign_registration_id_values is None:
        reassign_registration_id_values = set()
    else:
        reassign_registration_id_values = set(reassign_registration_id_values)

    normalized_primary_registration_id = str(primary_registration_id or '').strip()
    if not normalized_primary_registration_id:
        raise ValueError('primary_registration_id is required')

    password = secrets.token_urlsafe(12)[:12]
    temp_password_payload = signing.dumps({"p": password, "generated_at": timezone.now().isoformat()})
    normalized_first_name = str(first_name or '').strip()
    normalized_last_name = str(last_name or '').strip()
    if not normalized_first_name or not normalized_last_name:
        raise ValueError('first_name and last_name are required to provision a student account')

    if existing_user is not None:
        with transaction.atomic():
            user = existing_user
            user.set_password(password)
            user.must_reset_password = True
            user.temporary_password_encrypted = temp_password_payload
            user.first_name = normalized_first_name
            user.last_name = normalized_last_name
            user.student_class = student_class_int
            if normalized_gender is not None:
                user.gender = normalized_gender
            elif user.gender not in {'BOYS', 'GIRLS'}:
                user.gender = None
            user.school_category_extra = section
            user.is_active = True
            user.approval_status = 'approved'

            current_registration_id = str(getattr(user, 'registration_id', '') or '')
            if (not current_registration_id) or (current_registration_id in reassign_registration_id_values):
                user.registration_id = normalized_primary_registration_id

            user.save()
            username = user.username
    else:
        last_error = None
        for _ in range(5):
            username = _generate_unique_student_username(username_seed, participant_school.id)
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        email=_student_placeholder_email(username),
                        first_name=normalized_first_name,
                        last_name=normalized_last_name,
                        role='student',
                        student_class=student_class_int,
                        gender=normalized_gender,
                        school_category_extra=section,
                        school=participant_school,
                        registration_id=normalized_primary_registration_id,
                        is_active=True,
                        approval_status='approved',
                        must_reset_password=True,
                        temporary_password_encrypted=temp_password_payload,
                    )
                last_error = None
                break
            except IntegrityError as exc:
                last_error = exc
                continue

        if last_error is not None:
            raise last_error

    return {
        'user': user,
        'username': username,
        'password': password,
        'section': section,
    }


SCHOOL_LEVEL_CHOICES = {'LP', 'UP', 'HS', 'HSS'}
SCHOOL_GENDER_CHOICES = {'BOYS', 'GIRLS', 'MIXED'}
PARTICIPATION_TYPE_CHOICES = {'INDIVIDUAL', 'GROUP'}

GROUP_CLASS_CHOICES = SCHOOL_LEVEL_CHOICES
GROUP_GENDER_CHOICES = SCHOOL_GENDER_CHOICES


def _group_events_queryset():
    from events.models import Event
    return Event.objects.select_related(
        'event_definition',
        'event_definition__participation_type',
    ).filter(
        event_definition__participation_type__type_name='GROUP'
    )


def _individual_events_queryset():
    from events.models import Event
    return Event.objects.select_related(
        'event_definition',
        'event_definition__participation_type',
    ).filter(
        event_definition__participation_type__type_name='INDIVIDUAL'
    )


def _student_class_to_level_code(student_class):
    try:
        student_class_int = int(student_class)
    except (TypeError, ValueError):
        raise ValueError('student_class must be a valid integer')

    if 1 <= student_class_int <= 3:
        return 'LP'
    if 4 <= student_class_int <= 7:
        return 'UP'
    if 8 <= student_class_int <= 10:
        return 'HS'
    if 11 <= student_class_int <= 12:
        return 'HSS'
    raise ValueError('student_class must be between 1 and 12')


def _resolve_school_level_code(student_class=None, level_code=None):
    resolved_from_class = None
    if student_class not in [None, '']:
        resolved_from_class = _student_class_to_level_code(student_class)

    normalized_level_code = str(level_code or '').strip().upper()
    if normalized_level_code:
        if normalized_level_code not in SCHOOL_LEVEL_CHOICES:
            raise ValueError('level_code must be one of LP, UP, HS, HSS')
        if resolved_from_class and resolved_from_class != normalized_level_code:
            raise ValueError(
                f'level_code {normalized_level_code} does not match student_class-derived level {resolved_from_class}'
            )
        return normalized_level_code

    if resolved_from_class:
        return resolved_from_class

    raise ValueError('student_class or level_code is required')


def _normalize_person_name(first_name, last_name):
    return f"{(first_name or '').strip()} {(last_name or '').strip()}".strip()


def _resolve_group_leader_user(school_user, leader_first_name, leader_last_name):
    participant_school = getattr(school_user, 'school', None)
    if participant_school is None:
        participant_school = _ensure_school_user_linked_to_school_profile(school_user)
    if participant_school is None:
        return None

    return User.objects.filter(
        role='student',
        school=participant_school,
        first_name__iexact=(leader_first_name or '').strip(),
        last_name__iexact=(leader_last_name or '').strip(),
        approval_status='approved',
        is_active=True,
    ).order_by('-id').first()


def _parse_event_ids(event_ids):
    if event_ids in [None, '']:
        return []

    if isinstance(event_ids, str):
        tokens = [t.strip() for t in event_ids.split(',') if t.strip()]
        parsed = []
        for token in tokens:
            try:
                parsed.append(int(token))
            except (TypeError, ValueError):
                raise ValueError(f'Invalid event ID format: {token}')
        return parsed

    if not isinstance(event_ids, list):
        raise ValueError('event_ids must be a list of integers')

    parsed = []
    for value in event_ids:
        try:
            parsed.append(int(value))
        except (TypeError, ValueError):
            raise ValueError(f'Invalid event ID format: {value}')
    return parsed


def _eligible_school_events_by_matrix(
    events_queryset,
    required_participation_type,
    level_code,
    gender_category,
    include_metadata=False,
):
    from catalog.models import EventRule

    normalized_participation_type = str(required_participation_type or '').strip().upper()
    normalized_level_code = str(level_code or '').strip().upper()
    normalized_gender_category = str(gender_category or '').strip().upper()

    if normalized_participation_type not in PARTICIPATION_TYPE_CHOICES:
        raise ValueError('required_participation_type must be INDIVIDUAL or GROUP')
    if normalized_level_code not in SCHOOL_LEVEL_CHOICES:
        raise ValueError('level_code must be one of LP, UP, HS, HSS')
    if normalized_gender_category not in SCHOOL_GENDER_CHOICES:
        raise ValueError('gender_category must be one of BOYS, GIRLS, MIXED')

    if hasattr(events_queryset, 'select_related'):
        events = list(
            events_queryset.select_related(
                'event_definition',
                'event_definition__participation_type',
            )
        )
    else:
        events = list(events_queryset)
    if not events:
        if include_metadata:
            return set(), {}
        return set()

    eligible_event_ids = set()
    eligibility_metadata = {} if include_metadata else None
    candidate_events = []

    for event in events:
        event_id = getattr(event, 'id', None)
        event_definition_id = getattr(event, 'event_definition_id', None)
        event_variant_id = getattr(event, 'event_variant_id', None)

        participation_type_name = ''
        if event_definition_id:
            participation_type_name = str(
                getattr(
                    getattr(getattr(event, 'event_definition', None), 'participation_type', None),
                    'type_name',
                    '',
                ) or ''
            ).strip().upper()

        reason = None
        if not event_definition_id:
            reason = 'missing_event_definition'
        elif participation_type_name != normalized_participation_type:
            reason = 'participation_type_mismatch'
        elif str(getattr(event, 'status', '') or '').strip().lower() != 'published':
            reason = 'event_not_published'
        else:
            candidate_events.append(event)

        if include_metadata and event_id is not None:
            eligibility_metadata[event_id] = {
                'eligible': False,
                'reason': reason,
                'event_definition_id': event_definition_id,
                'event_variant_id': event_variant_id,
                'participation_type': participation_type_name,
                'selected_rule_variant_id': None,
                'selected_rule_genders': [],
                'rule_source': 'none',
            }

    definition_ids = {
        event.event_definition_id
        for event in candidate_events
        if getattr(event, 'event_definition_id', None)
    }
    if not definition_ids:
        if include_metadata:
            return set(), eligibility_metadata
        return set()

    rules = EventRule.objects.filter(
        event_id__in=definition_ids,
        level__level_code=normalized_level_code,
    ).values('event_id', 'variant_id', 'gender_eligibility')

    rules_by_event_definition = {}
    for rule in rules:
        rules_by_event_definition.setdefault(rule['event_id'], []).append(rule)

    for event in candidate_events:
        event_rules = rules_by_event_definition.get(event.event_definition_id, [])
        rule_source = 'none'
        selected_rules = []
        if event.event_variant_id:
            variant_rules = [rule for rule in event_rules if rule['variant_id'] == event.event_variant_id]
            if variant_rules:
                selected_rules = variant_rules
                rule_source = 'variant_specific'
            else:
                selected_rules = [rule for rule in event_rules if rule['variant_id'] is None]
                if selected_rules:
                    rule_source = 'default_variant'
        else:
            selected_rules = [rule for rule in event_rules if rule['variant_id'] is None]
            if selected_rules:
                rule_source = 'default_variant'
            else:
                variant_pool_rules = [rule for rule in event_rules if rule['variant_id'] is not None]
                if variant_pool_rules:
                    selected_rules = variant_pool_rules
                    rule_source = 'variant_pool_fallback'

        eligible_genders = {rule['gender_eligibility'] for rule in selected_rules}
        is_eligible = normalized_gender_category in eligible_genders
        if (
            not is_eligible
            and normalized_participation_type == 'INDIVIDUAL'
            and normalized_gender_category in {'BOYS', 'GIRLS'}
            and 'MIXED' in eligible_genders
        ):
            is_eligible = True
        if is_eligible:
            eligible_event_ids.add(event.id)

        if include_metadata:
            eligibility_metadata[event.id].update({
                'eligible': is_eligible,
                'reason': (
                    'eligible'
                    if is_eligible
                    else ('no_level_rule' if not selected_rules else 'gender_mismatch')
                ),
                'selected_rule_variant_id': selected_rules[0]['variant_id'] if selected_rules else None,
                'selected_rule_genders': sorted(eligible_genders),
                'rule_source': rule_source,
            })

    if include_metadata:
        return eligible_event_ids, eligibility_metadata
    return eligible_event_ids


def _group_events_eligible_for_class_and_gender(
    events_queryset,
    group_class,
    gender_category,
    include_metadata=False,
):
    return _eligible_school_events_by_matrix(
        events_queryset=events_queryset,
        required_participation_type='GROUP',
        level_code=group_class,
        gender_category=gender_category,
        include_metadata=include_metadata,
    )


def _individual_events_eligible_for_class_and_gender(
    events_queryset,
    level_code,
    gender_category,
    include_metadata=False,
):
    return _eligible_school_events_by_matrix(
        events_queryset=events_queryset,
        required_participation_type='INDIVIDUAL',
        level_code=level_code,
        gender_category=gender_category,
        include_metadata=include_metadata,
    )


def _validate_group_event_ids(event_ids, group_class=None, gender_category=None):
    parsed_ids = _parse_event_ids(event_ids)
    if not parsed_ids:
        return []

    valid_events_queryset = _group_events_queryset().filter(id__in=parsed_ids)
    valid_ids = set(valid_events_queryset.values_list('id', flat=True))
    invalid_ids = sorted(set(parsed_ids) - valid_ids)
    if invalid_ids:
        raise ValueError(f'Invalid group event IDs: {invalid_ids}')

    normalized_group_class = str(group_class or '').strip().upper()
    normalized_gender_category = str(gender_category or '').strip().upper()
    if normalized_group_class or normalized_gender_category:
        if normalized_group_class not in GROUP_CLASS_CHOICES:
            raise ValueError('group_class must be one of LP, UP, HS, HSS')
        if normalized_gender_category not in GROUP_GENDER_CHOICES:
            raise ValueError('gender_category must be one of BOYS, GIRLS, MIXED')

        eligible_ids = _group_events_eligible_for_class_and_gender(
            events_queryset=valid_events_queryset,
            group_class=normalized_group_class,
            gender_category=normalized_gender_category,
        )
        ineligible_ids = sorted(valid_ids - eligible_ids)
        if ineligible_ids:
            raise ValueError(
                f'Ineligible group event IDs for group_class {normalized_group_class} '
                f'and gender_category {normalized_gender_category}: {ineligible_ids}'
            )

    return sorted(valid_ids)


def _validate_individual_event_ids(event_ids, level_code=None, gender_category=None):
    parsed_ids = _parse_event_ids(event_ids)
    if not parsed_ids:
        return []

    valid_events_queryset = _individual_events_queryset().filter(id__in=parsed_ids)
    valid_ids = set(valid_events_queryset.values_list('id', flat=True))
    invalid_ids = sorted(set(parsed_ids) - valid_ids)
    if invalid_ids:
        raise ValueError(f'Invalid individual event IDs: {invalid_ids}')

    normalized_level_code = str(level_code or '').strip().upper()
    normalized_gender_category = str(gender_category or '').strip().upper()
    if normalized_level_code or normalized_gender_category:
        if normalized_level_code not in SCHOOL_LEVEL_CHOICES:
            raise ValueError('level_code must be one of LP, UP, HS, HSS')
        if normalized_gender_category not in SCHOOL_GENDER_CHOICES:
            raise ValueError('gender_category must be one of BOYS, GIRLS, MIXED')

        eligible_ids = _individual_events_eligible_for_class_and_gender(
            events_queryset=valid_events_queryset,
            level_code=normalized_level_code,
            gender_category=normalized_gender_category,
        )
        ineligible_ids = sorted(valid_ids - eligible_ids)
        if ineligible_ids:
            raise ValueError(
                f'Ineligible individual event IDs for level_code {normalized_level_code} '
                f'and gender_category {normalized_gender_category}: {ineligible_ids}'
            )

    return sorted(valid_ids)


def _normalize_member_payloads(participants):
    if not isinstance(participants, list) or not participants:
        raise ValueError('participants must be a non-empty list')

    normalized = []
    for idx, participant in enumerate(participants, start=1):
        if not isinstance(participant, dict):
            raise ValueError(f'Invalid participant format at position {idx}')

        first_name = str(participant.get('first_name') or '').strip()
        last_name = str(participant.get('last_name') or '').strip()
        if not first_name or not last_name:
            raise ValueError(f'First name and last name are required for participant at position {idx}')

        raw_gender = str(_read_compat_field(participant, 'gender', 'genderCategory') or '').strip().upper()
        gender = raw_gender if raw_gender in {'BOYS', 'GIRLS'} else None
        if raw_gender and gender is None:
            raise ValueError(f'gender must be BOYS or GIRLS for participant at position {idx}')

        student_class_raw = _read_compat_field(participant, 'student_class', 'studentClass')
        student_class = None
        if student_class_raw not in [None, '']:
            try:
                student_class = int(student_class_raw)
            except (TypeError, ValueError):
                raise ValueError(f'student_class must be a valid number for participant at position {idx}')
            if not (1 <= student_class <= 12):
                raise ValueError(f'student_class must be between 1 and 12 for participant at position {idx}')

        raw_phone = str(_read_compat_field(participant, 'phone', 'phoneNumber') or '').strip()
        phone = ''
        if raw_phone:
            phone_digits = ''.join(ch for ch in raw_phone if ch.isdigit())
            if len(phone_digits) != 10:
                raise ValueError(f'Phone number must be exactly 10 digits for participant at position {idx}')
            if phone_digits[0] not in {'7', '8', '9'}:
                raise ValueError(f'Phone number must start with 7, 8, or 9 for participant at position {idx}')
            if phone_digits == '0000000000':
                raise ValueError(f'Phone number cannot be all zeros for participant at position {idx}')
            phone = phone_digits

        normalized.append({
            'member_order': idx,
            'first_name': first_name,
            'last_name': last_name,
            'gender': gender,
            'student_class': student_class,
            'phone': phone,
        })

    return normalized


def _resolve_leader_member(normalized_members, leader_index=None, leader_name=None):
    leader_member = None

    if leader_index not in [None, '']:
        try:
            leader_index_int = int(leader_index)
        except (TypeError, ValueError):
            raise ValueError('leader_index must be a valid integer')

        if leader_index_int < 1 or leader_index_int > len(normalized_members):
            raise ValueError('leader_index is out of range for participants')
        leader_member = normalized_members[leader_index_int - 1]

    if leader_name not in [None, '']:
        leader_name_normalized = str(leader_name).strip().lower()
        matches = [
            member for member in normalized_members
            if _normalize_person_name(member['first_name'], member['last_name']).lower() == leader_name_normalized
        ]
        if not matches:
            raise ValueError('leader_name must match one of the participants')

        if leader_member is not None and leader_member != matches[0]:
            raise ValueError('leader_index and leader_name refer to different participants')

        leader_member = matches[0]

    if leader_member is None:
        raise ValueError('Either leader_index or leader_name is required')

    return leader_member


def _read_compat_field(data, snake_key, camel_key):
    snake_value = data.get(snake_key)
    if snake_value not in [None, '']:
        return snake_value

    camel_value = data.get(camel_key)
    if camel_value not in [None, '']:
        return camel_value

    if snake_key in data:
        return snake_value
    return camel_value


def _create_school_group_entry_from_payload(school_user, payload, source='manual'):
    group_id = str(payload.get('group_id') or '').strip().upper()
    group_class = str(_read_compat_field(payload, 'group_class', 'groupClass') or '').strip().upper()
    gender_category = str(_read_compat_field(payload, 'gender_category', 'genderCategory') or '').strip().upper()

    participant_count_raw = _read_compat_field(payload, 'participant_count', 'participantCount')
    try:
        participant_count = int(participant_count_raw)
    except (TypeError, ValueError):
        raise ValueError('participant_count must be a valid number')

    if participant_count < 1 or participant_count > 20:
        raise ValueError('participant_count must be between 1 and 20')

    if not group_id:
        raise ValueError('group_id is required')
    if group_class not in GROUP_CLASS_CHOICES:
        raise ValueError('group_class must be one of LP, UP, HS, HSS')
    if gender_category not in GROUP_GENDER_CHOICES:
        raise ValueError('gender_category must be one of BOYS, GIRLS, MIXED')

    if SchoolGroupEntry.objects.filter(school=school_user, group_id__iexact=group_id).exists():
        raise ValueError(f'group_id already exists: {group_id}')

    normalized_members = _normalize_member_payloads(payload.get('participants') or [])
    if len(normalized_members) != participant_count:
        raise ValueError('participant_count must match the number of participants provided')

    leader_member = _resolve_leader_member(
        normalized_members=normalized_members,
        leader_index=_read_compat_field(payload, 'leader_index', 'leaderIndex'),
        leader_name=_read_compat_field(payload, 'leader_name', 'leaderName'),
    )
    leader_full_name = _normalize_person_name(leader_member['first_name'], leader_member['last_name'])

    valid_event_ids = _validate_group_event_ids(
        _read_compat_field(payload, 'event_ids', 'eventIds'),
        group_class=group_class,
        gender_category=gender_category,
    )

    with transaction.atomic():
        group_entry = SchoolGroupEntry.objects.create(
            school=school_user,
            group_id=group_id,
            group_class=group_class,
            gender_category=gender_category,
            participant_count=participant_count,
            leader_full_name=leader_full_name,
            leader_user=_resolve_group_leader_user(
                school_user,
                leader_member['first_name'],
                leader_member['last_name'],
            ),
            source=source,
            status='pending',
        )

        for member in normalized_members:
            SchoolGroupMember.objects.create(
                group_entry=group_entry,
                member_order=member['member_order'],
                first_name=member['first_name'],
                last_name=member['last_name'],
                gender=member['gender'],
                student_class=member['student_class'],
                phone=member['phone'],
                is_leader=(member['member_order'] == leader_member['member_order']),
            )

        if valid_event_ids:
            group_entry.events.set(valid_event_ids)

    return group_entry


def _parse_group_members_from_import_names(raw_names):
    text = str(raw_names or '').replace('\n', '|').replace(';', '|').strip()
    if not text:
        raise ValueError('participant names are required')

    tokens = [token.strip() for token in text.split('|') if token.strip()]
    participants = []
    for token in tokens:
        parts = token.split()
        if len(parts) < 2:
            raise ValueError(f'Invalid participant name "{token}". Use "FirstName LastName".')
        participants.append({
            'first_name': parts[0],
            'last_name': ' '.join(parts[1:]),
        })
    return participants


def _build_group_payload_from_import_row(row):
    normalized_row = {}
    for key, value in row.items():
        normalized_key = re.sub(r'[^a-z0-9]+', '_', str(key or '').strip().lower()).strip('_')
        normalized_row[normalized_key] = value

    participants = _parse_group_members_from_import_names(
        normalized_row.get('participant_names') or normalized_row.get('participants')
    )

    return {
        'group_id': normalized_row.get('group_id'),
        'group_class': normalized_row.get('group_class'),
        'gender_category': normalized_row.get('gender') or normalized_row.get('gender_category'),
        'participant_count': normalized_row.get('number_of_participants') or normalized_row.get('participant_count'),
        'participants': participants,
        'leader_name': normalized_row.get('leader_name'),
        'event_ids': normalized_row.get('event_ids') or normalized_row.get('events'),
    }


# Admin: Create School Accounts
class AdminCreateSchoolView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    
    def create(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        email = request.data.get('email')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        school_model_id = request.data.get('school_model_id')  # ID from School model
        
        if not all([username, password, email]):
            return Response({'error': 'Username, password, and email are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Username already exists'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if User.objects.filter(email=email).exists():
            return Response({'error': 'Email already exists'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Create school user
        school_user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name or '',
            last_name=last_name or '',
            role='school',
            is_active=True
        )
        
        # Link to School model if provided
        if school_model_id:
            from .models import School
            try:
                school_obj = School.objects.get(id=school_model_id)
                school_user.school = school_obj
                school_user.save()
            except School.DoesNotExist:
                pass
        
        # Send email with credentials
        try:
            frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:3000')
            subject = 'School Account Created - E-Kalolsavam'
            message = f'''Your school account has been created.
            
Username: {username}
Password: {password}

Please log in at {frontend_url}/login and change your password for security.
            
You can now submit participant data for your school.'''
            send_mail(subject, message, 'joelfrancisjoy@gmail.com', [email], fail_silently=True)
        except Exception:
            pass
        
        return Response({'message': 'School account created successfully', 
                        'username': username}, status=status.HTTP_201_CREATED)


# Admin: Generate IDs for Volunteers and Judges with name assignments
class AdminGenerateIDView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    
    def create(self, request):
        role = request.data.get('role')  # 'student', 'volunteer' or 'judge'
        count = request.data.get('count', 1)
        assignments = request.data.get('assignments', [])  # List of {name, phone, student_class, school_id} dicts
        
        if role not in ['student', 'volunteer', 'judge']:
            return Response({'error': 'Invalid role. Must be student, volunteer or judge'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # If assignments provided, use that count; otherwise use count parameter
        if assignments:
            if not isinstance(assignments, list):
                return Response({'error': 'Assignments must be a list'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            count = len(assignments)
        else:
            try:
                count = int(count)
                if count < 1 or count > 100:
                    return Response({'error': 'Count must be between 1 and 100'}, 
                                  status=status.HTTP_400_BAD_REQUEST)
            except (ValueError, TypeError):
                return Response({'error': 'Invalid count value'}, 
                              status=status.HTTP_400_BAD_REQUEST)
        
        ids = []
        # Generate prefix based on role
        if role == 'volunteer':
            prefix = 'VOL'
        elif role == 'judge':
            prefix = 'JUD'
        elif role == 'student':
            prefix = 'STU'
        else:
            return Response({'error': 'Invalid role. Must be student, volunteer, or judge'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        for i in range(count):
            # Get assignment details if provided
            assigned_name = None
            assigned_phone = None
            notes = ''
            
            if assignments and i < len(assignments):
                assignment = assignments[i]
                assigned_name = assignment.get('name', '').strip() or None
                assigned_phone = assignment.get('phone', '').strip() or None
                notes = assignment.get('notes', '').strip()
            
            # Generate unique 4-digit random number ID
            max_attempts = 1000
            attempt = 0
            id_code = None
            
            while attempt < max_attempts:
                # Generate random 4-digit number (1000-9999)
                random_number = secrets.randbelow(9000) + 1000
                id_code = f"{prefix}{random_number}"
                
                # Check if ID already exists
                if not AdminIssuedID.objects.filter(id_code=id_code).exists():
                    break
                    
                attempt += 1
                id_code = None
            
            if id_code is None:
                return Response({
                    'error': f'Unable to generate unique ID after {max_attempts} attempts. Please try again.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Create the ID with assignment details
            admin_id = AdminIssuedID.objects.create(
                id_code=id_code,
                role=role,
                assigned_name=assigned_name,
                assigned_phone=assigned_phone,
                notes=notes,
                is_active=True,
                created_by=request.user
            )
            ids.append(AdminIssuedIDSerializer(admin_id).data)
        
        return Response({
            'ids': ids,
            'count': len(ids),
            'message': f'Successfully generated {len(ids)} {role} ID(s)'
        }, status=status.HTTP_201_CREATED)


# Public: Sign up with ID (for volunteers and judges)
class IDSignupView(generics.CreateAPIView):
    permission_classes = [AllowAny]
    
    def create(self, request):
        id_code = request.data.get('id_code')
        username = request.data.get('username')
        password = request.data.get('password')
        email = request.data.get('email')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        phone = request.data.get('phone')
        
        if not all([id_code, username, password, email, first_name, last_name]):
            return Response({'error': 'All fields are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Validate ID
        try:
            issued_id = AdminIssuedID.objects.get(id_code=id_code.strip().upper())
        except AdminIssuedID.DoesNotExist:
            return Response({'error': 'Invalid ID code. Please check and try again.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Check if ID is active
        if not issued_id.is_active:
            return Response({'error': 'This ID has been deactivated. Please contact admin.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Check if ID is already used
        if issued_id.is_used:
            return Response({'error': 'This ID has already been used for registration.'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Verify name matches (if name was pre-assigned)
        if issued_id.assigned_name:
            provided_name = f"{first_name} {last_name}".strip().lower()
            assigned_name = issued_id.assigned_name.strip().lower()
            
            # Allow for minor variations but check substantial match
            if provided_name != assigned_name and assigned_name not in provided_name and provided_name not in assigned_name:
                return Response({
                    'error': f'Name mismatch. This ID is assigned to: {issued_id.assigned_name}. Please verify your details.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify phone matches (if phone was pre-assigned)
        if issued_id.assigned_phone and phone:
            # Normalize phone numbers for comparison
            provided_phone = ''.join(filter(str.isdigit, phone))
            assigned_phone = ''.join(filter(str.isdigit, issued_id.assigned_phone))
            
            if provided_phone != assigned_phone:
                return Response({
                    'error': 'Phone number does not match our records. Please verify your details.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if username or email already exists
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Username already exists'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if User.objects.filter(email=email).exists():
            return Response({'error': 'Email already exists'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Create user with inactive status (requires admin verification)
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
            phone=phone or '',
            role=issued_id.role,
            approval_status='pending',
            is_active=False,  # Account inactive until admin verifies
            registration_id=id_code
        )
        
        # Mark ID as used
        issued_id.is_used = True
        issued_id.used_by = user
        from django.utils import timezone
        issued_id.used_at = timezone.now()
        issued_id.save()
        
        # Create signup request for admin review
        IDSignupRequest.objects.create(
            issued_id=issued_id,
            user=user,
            status='pending'
        )
        
        # Send email notification to user
        try:
            subject = f'{issued_id.role.title()} Registration Received - E-Kalolsavam'
            message = f'''Dear {first_name},

Your {issued_id.role} registration has been received with ID: {id_code}

Your account is currently pending admin verification. You will receive an email once your account is activated.

Username: {username}
Role: {issued_id.role.title()}

Please do not share your login credentials with anyone.

Thank you for joining E-Kalolsavam!
'''
            send_mail(subject, message, 'joelfrancisjoy@gmail.com', [email], fail_silently=True)
        except Exception:
            pass
        
        return Response({
            'message': f'Registration successful! Your {issued_id.role} account is pending admin verification. You will be notified via email once approved.',
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'status': 'pending_verification'
            }
        }, status=status.HTTP_201_CREATED)


# Admin: List and approve signup requests
class IDSignupRequestListView(generics.ListAPIView):
    serializer_class = IDSignupRequestSerializer
    permission_classes = [IsAuthenticated, IsAdminOrVolunteerRole]
    
    def get_queryset(self):
        status_filter = self.request.query_params.get('status', 'pending')
        
        # If status is 'all', return all requests regardless of status
        if status_filter == 'all':
            return IDSignupRequest.objects.all().order_by('-requested_at')
        else:
            return IDSignupRequest.objects.filter(status=status_filter).order_by('-requested_at')


class IDSignupRequestDetailView(generics.RetrieveUpdateAPIView):
    queryset = IDSignupRequest.objects.all()
    serializer_class = IDSignupRequestSerializer
    permission_classes = [IsAuthenticated, IsAdminOrVolunteerRole]
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        new_status = request.data.get('status')
        
        if new_status not in ['approved', 'rejected']:
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if the current user is a volunteer trying to approve non-student requests
        user = request.user
        if user.role == 'volunteer':
            # Volunteers can only approve student signup requests
            if instance.issued_id.role != 'student':
                return Response(
                    {'error': 'Volunteers can only approve student registration requests'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        instance.status = new_status
        instance.reviewed_by = request.user
        from django.utils import timezone
        instance.reviewed_at = timezone.now()
        instance.notes = request.data.get('notes', '')
        instance.save()
        
        # Get the associated ID and user
        issued_id = instance.issued_id
        user_obj = instance.user
        
        # If approved, activate the user and mark ID as verified
        if new_status == 'approved':
            user_obj.approval_status = 'approved'
            user_obj.is_active = True
            user_obj.save()
            
            # Mark the ID as verified
            issued_id.is_verified = True
            issued_id.verified_by = request.user
            issued_id.verified_at = timezone.now()
            issued_id.save()
            
            # Send approval email
            try:
                subject = f'{user_obj.role.title()} Account Approved - E-Kalolsavam'
                message = f'''Dear {user_obj.first_name},

Congratulations! Your {user_obj.role} account has been approved and activated.

You can now log in to the E-Kalolsavam platform using your credentials:
Username: {user_obj.username}

Please keep your login credentials secure and do not share them with anyone.

Welcome to the E-Kalolsavam team!

Best regards,
E-Kalolsavam Admin Team
'''
                send_mail(subject, message, 'joelfrancisjoy@gmail.com', [user_obj.email], fail_silently=True)
            except Exception:
                pass
                
        elif new_status == 'rejected':
            user_obj.approval_status = 'rejected'
            user_obj.is_active = False
            user_obj.save()
            
            # Optionally, free up the ID for reuse
            if request.data.get('free_id', False):
                issued_id.is_used = False
                issued_id.used_by = None
                issued_id.used_at = None
                issued_id.save()
            
            # Send rejection email
            try:
                subject = f'{user_obj.role.title()} Registration Update - E-Kalolsavam'
                rejection_reason = instance.notes or 'verification failed'
                message = f'''Dear {user_obj.first_name},

Thank you for your interest in joining E-Kalolsavam as a {user_obj.role}.

After review, we regret to inform you that your registration could not be approved at this time.
Reason: {rejection_reason}

If you believe this is an error or have questions, please contact the admin team.

Thank you for your understanding.

Best regards,
E-Kalolsavam Admin Team
'''
                send_mail(subject, message, 'joelfrancisjoy@gmail.com', [user_obj.email], fail_silently=True)
            except Exception:
                pass
        
        return Response({
            'message': f'Request {new_status} successfully',
            'request': IDSignupRequestSerializer(instance).data
        })


# School: Submit participant data
class SchoolSubmitParticipantsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Check if user has school role
        if not hasattr(request.user, 'role') or request.user.role != 'school':
            return Response(
                {'error': 'Only school accounts can submit participants'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        participants_data = request.data.get('participants', [])
        if not isinstance(participants_data, list) or not participants_data:
            return Response({'error': 'participants must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        errors = []
        payload_participant_ids = set()
        for idx, p_data in enumerate(participants_data, start=1):
            if not isinstance(p_data, dict):
                errors.append({'row': idx, 'participant_id': None, 'error': f'Invalid participant payload format at index {idx}'})
                continue

            # Validate required fields
            participant_id = _read_compat_field(p_data, 'participant_id', 'participantId')
            first_name = _read_compat_field(p_data, 'first_name', 'firstName')
            last_name = _read_compat_field(p_data, 'last_name', 'lastName')
            student_class = _read_compat_field(p_data, 'student_class', 'studentClass')

            gender_value = p_data.get('gender')
            if gender_value in [None, '']:
                gender_value = _read_compat_field(p_data, 'gender_category', 'genderCategory')
            gender = str(gender_value or '').strip().upper()

            normalized_participant_id = str(participant_id or '').strip()
            normalized_participant_id_upper = normalized_participant_id.upper()
            if normalized_participant_id_upper:
                if normalized_participant_id_upper in payload_participant_ids:
                    errors.append({
                        'row': idx,
                        'participant_id': normalized_participant_id,
                        'error': f'Duplicate participant_id in request payload: {normalized_participant_id_upper}',
                    })
                    continue
                payload_participant_ids.add(normalized_participant_id_upper)
            
            if not all([participant_id, first_name, last_name, student_class, gender]):
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id or None,
                    'error': f'Missing required fields for participant {participant_id or "unknown"}',
                })
                continue

            if gender not in ['BOYS', 'GIRLS']:
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id or None,
                    'error': f'Invalid gender for participant {participant_id}. Must be BOYS or GIRLS',
                })
                continue
            
            # Validate student_class is between 1 and 12
            try:
                student_class_int = int(student_class)
            except (ValueError, TypeError):
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id or None,
                    'error': f'Invalid student class for participant {participant_id}',
                })
                continue

            try:
                participant_level_code = _student_class_to_level_code(student_class_int)
            except ValueError:
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id or None,
                    'error': f'Student class must be between 1 and 12 for participant {participant_id}',
                })
                continue

            event_ids = _read_compat_field(p_data, 'event_ids', 'eventIds')
            if event_ids in [None, '']:
                event_ids = []
            try:
                valid_event_ids = _validate_individual_event_ids(
                    event_ids,
                    level_code=participant_level_code,
                    gender_category=gender,
                )
            except ValueError as exc:
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id or None,
                    'error': f'{exc} for participant {participant_id}',
                })
                continue

            if SchoolParticipant.objects.filter(
                school=request.user,
                participant_id__iexact=normalized_participant_id,
            ).exists():
                errors.append({
                    'row': idx,
                    'participant_id': normalized_participant_id,
                    'error': f'participant_id already exists: {normalized_participant_id_upper}',
                })
                continue

            with transaction.atomic():
                participant = SchoolParticipant.objects.create(
                    school=request.user,
                    participant_id=normalized_participant_id,
                    first_name=first_name,
                    last_name=last_name,
                    student_class=student_class_int,
                    gender=gender,
                )

                if valid_event_ids:
                    participant.events.set(valid_event_ids)
            
            created.append(SchoolParticipantSerializer(participant).data)

        status_code = status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST
        return Response({
            'participants': created,
            'count': len(created),
            'error_count': len(errors),
            'errors': errors,
        }, status=status_code)


class SchoolSubmitGroupParticipantsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if getattr(request.user, 'role', None) != 'school':
            return Response(
                {'error': 'Only school accounts can submit group participants'},
                status=status.HTTP_403_FORBIDDEN
            )

        groups_data = request.data.get('groups', [])
        if not isinstance(groups_data, list) or not groups_data:
            return Response({'error': 'groups must be a non-empty list'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        errors = []
        payload_group_ids = set()

        for idx, group_payload in enumerate(groups_data, start=1):
            if not isinstance(group_payload, dict):
                errors.append({'row': idx, 'group_id': None, 'error': f'Group #{idx}: Invalid group payload format at index {idx}'})
                continue

            candidate_group_id = str(group_payload.get('group_id') or '').strip().upper()
            if candidate_group_id:
                if candidate_group_id in payload_group_ids:
                    errors.append({
                        'row': idx,
                        'group_id': candidate_group_id,
                        'error': f'Group #{idx}: Duplicate group_id in request payload: {candidate_group_id}',
                    })
                    continue
                payload_group_ids.add(candidate_group_id)

            try:
                group_entry = _create_school_group_entry_from_payload(
                    school_user=request.user,
                    payload=group_payload,
                    source='manual',
                )
            except ValueError as exc:
                errors.append({
                    'row': idx,
                    'group_id': candidate_group_id or None,
                    'error': f'Group #{idx}: {exc}',
                })
                continue

            created.append(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data)

        status_code = status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST
        return Response({
            'groups': created,
            'count': len(created),
            'error_count': len(errors),
            'errors': errors,
        }, status=status_code)


class SchoolViewOwnGroupParticipantsView(generics.ListAPIView):
    serializer_class = SchoolGroupEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if getattr(self.request.user, 'role', None) != 'school':
            return SchoolGroupEntry.objects.none()

        queryset = SchoolGroupEntry.objects.filter(school=self.request.user).select_related(
            'leader_user',
            'reviewed_by',
        ).prefetch_related(
            'members',
            'events',
        ).order_by('-submitted_at')

        status_filter = (self.request.query_params.get('status') or '').strip().lower()
        if status_filter in {'pending', 'approved', 'rejected'}:
            queryset = queryset.filter(status=status_filter)

        group_class = (self.request.query_params.get('group_class') or '').strip().upper()
        if group_class in GROUP_CLASS_CHOICES:
            queryset = queryset.filter(group_class=group_class)

        return queryset


@extend_schema(
    methods=['GET'],
    parameters=[SchoolIndividualEventsQuerySerializer],
    responses={
        200: EventSerializer(many=True),
        400: OpenApiResponse(response=ErrorEnvelopeSerializer),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def school_individual_events(request):
    if getattr(request.user, 'role', None) != 'school':
        return _error_response(
            'Only school accounts can access this endpoint',
            status.HTTP_403_FORBIDDEN,
            code='forbidden',
        )

    student_class_raw = _read_compat_field(request.query_params, 'student_class', 'studentClass')
    level_code_raw = _read_compat_field(request.query_params, 'level_code', 'levelCode')
    gender_category = str(
        _read_compat_field(request.query_params, 'gender_category', 'genderCategory') or ''
    ).strip().upper()

    queryset = _individual_events_queryset().filter(status='published')

    has_level_filter = student_class_raw not in [None, ''] or level_code_raw not in [None, '']
    has_gender_filter = bool(gender_category)

    if not has_level_filter and not has_gender_filter:
        queryset = queryset.order_by('date', 'start_time')
        return Response(EventSerializer(queryset, many=True).data)

    if has_level_filter != has_gender_filter:
        return _error_response(
            'student_class or level_code and gender_category must be provided together',
            status.HTTP_400_BAD_REQUEST,
            code='invalid_filters',
        )

    try:
        level_code = _resolve_school_level_code(student_class=student_class_raw, level_code=level_code_raw)
    except ValueError as exc:
        return _error_response(str(exc), status.HTTP_400_BAD_REQUEST, code='invalid_level')

    if gender_category not in SCHOOL_GENDER_CHOICES:
        return _error_response(
            'gender_category must be one of BOYS, GIRLS, MIXED',
            status.HTTP_400_BAD_REQUEST,
            code='invalid_gender_category',
        )

    eligible_event_ids = _individual_events_eligible_for_class_and_gender(
        events_queryset=queryset,
        level_code=level_code,
        gender_category=gender_category,
    )
    queryset = queryset.filter(id__in=eligible_event_ids).order_by('date', 'start_time')

    return Response(EventSerializer(queryset, many=True).data)


@extend_schema(
    methods=['GET'],
    parameters=[SchoolGroupEventsQuerySerializer],
    responses={
        200: EventSerializer(many=True),
        400: OpenApiResponse(response=ErrorEnvelopeSerializer),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def school_group_events(request):
    if getattr(request.user, 'role', None) != 'school':
        return _error_response(
            'Only school accounts can access this endpoint',
            status.HTTP_403_FORBIDDEN,
            code='forbidden',
        )
    queryset = _group_events_queryset().filter(status='published')

    group_class = str(_read_compat_field(request.query_params, 'group_class', 'groupClass') or '').strip().upper()
    gender_category = str(_read_compat_field(request.query_params, 'gender_category', 'genderCategory') or '').strip().upper()

    # If both missing -> return all published group events
    if not group_class and not gender_category:
        queryset = queryset.order_by('date', 'start_time')
        return Response(EventSerializer(queryset, many=True).data)

    # If one missing -> return 400
    if bool(group_class) != bool(gender_category):
        return _error_response(
            'group_class and gender_category must be provided together',
            status.HTTP_400_BAD_REQUEST,
            code='invalid_filters',
        )
    if group_class not in GROUP_CLASS_CHOICES:
        return _error_response(
            'group_class must be one of LP, UP, HS, HSS',
            status.HTTP_400_BAD_REQUEST,
            code='invalid_group_class',
        )
    if gender_category not in GROUP_GENDER_CHOICES:
        return _error_response(
            'gender_category must be one of BOYS, GIRLS, MIXED',
            status.HTTP_400_BAD_REQUEST,
            code='invalid_gender_category',
        )

    eligible_event_ids = _group_events_eligible_for_class_and_gender(
        events_queryset=queryset,
        group_class=group_class,
        gender_category=gender_category,
    )
    queryset = queryset.filter(id__in=eligible_event_ids)

    queryset = queryset.order_by('date', 'start_time')
    return Response(EventSerializer(queryset, many=True).data)


class SchoolBulkImportGroupParticipantsView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if getattr(request.user, 'role', None) != 'school':
            return Response(
                {'error': 'Only school accounts can import group participants'},
                status=status.HTTP_403_FORBIDDEN
            )

        upload = request.FILES.get('file')
        if upload is None:
            return Response({'error': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        filename = (upload.name or '').strip().lower()
        try:
            if filename.endswith('.csv'):
                rows = self._parse_csv(upload)
            elif filename.endswith('.xlsx'):
                rows = self._parse_xlsx(upload)
            else:
                return Response({'error': 'Unsupported file format. Use .csv or .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        imported = []
        errors = []
        seen_group_ids = set()

        for row_number, row_data in rows:
            try:
                payload = _build_group_payload_from_import_row(row_data)
                candidate_group_id = str(payload.get('group_id') or '').strip().upper()
                if candidate_group_id:
                    if candidate_group_id in seen_group_ids:
                        raise ValueError(f'Duplicate group_id in file: {candidate_group_id}')
                    seen_group_ids.add(candidate_group_id)

                group_entry = _create_school_group_entry_from_payload(
                    school_user=request.user,
                    payload=payload,
                    source='bulk',
                )
                imported.append(SchoolGroupEntrySerializer(group_entry, context={'request': request}).data)
            except ValueError as exc:
                errors.append({'row': row_number, 'error': str(exc)})

        status_code = status.HTTP_201_CREATED if imported else status.HTTP_400_BAD_REQUEST
        return Response({
            'imported_count': len(imported),
            'error_count': len(errors),
            'imported_groups': imported,
            'errors': errors,
        }, status=status_code)

    def _parse_csv(self, upload):
        decoded = upload.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        if not reader.fieldnames:
            raise ValueError('CSV header row is required')

        rows = []
        for index, row in enumerate(reader, start=2):
            rows.append((index, row))
        return rows

    def _parse_xlsx(self, upload):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('Excel import requires openpyxl package')

        workbook = load_workbook(upload, read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            raise ValueError('Excel header row is required')

        normalized_headers = [str(header or '').strip() for header in headers]
        rows = []
        for index, values in enumerate(iterator, start=2):
            row_data = {}
            for col_idx, header in enumerate(normalized_headers):
                row_data[header] = values[col_idx] if col_idx < len(values) else None
            rows.append((index, row_data))
        return rows


# School: Generate Student IDs
class SchoolGenerateStudentIDView(generics.CreateAPIView):
    """
    Allows schools to generate student IDs for pre-registration.
    Students will use these IDs to sign up on the platform.
    """
    permission_classes = [IsAuthenticated]
    
    def create(self, request):
        if request.user.role != 'school':
            raise PermissionDenied("Only school accounts can generate student IDs")
        
        assignments = request.data.get('students', [])  # List of {name, phone, student_class} dicts
        
        if not assignments or not isinstance(assignments, list):
            return Response({'error': 'Students list is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        if len(assignments) > 100:
            return Response({'error': 'Maximum 100 students per request'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        ids = []
        prefix = 'STU'
        
        for assignment in assignments:
            assigned_name = assignment.get('name', '').strip()
            assigned_phone = assignment.get('phone', '').strip() or None
            student_class = assignment.get('student_class')
            notes = assignment.get('notes', '').strip()
            
            if not assigned_name:
                continue  # Skip entries without name
            
            # Add school info to notes
            school_note = f"School: {request.user.school.name if request.user.school else request.user.username}"
            if student_class:
                school_note += f" | Class: {student_class}"
            if notes:
                school_note += f" | {notes}"
            else:
                notes = school_note
            
            # Generate unique ID
            max_attempts = 1000
            attempt = 0
            id_code = None
            
            while attempt < max_attempts:
                random_number = secrets.randbelow(9000) + 1000
                id_code = f"{prefix}{random_number}"
                
                if not AdminIssuedID.objects.filter(id_code=id_code).exists():
                    break
                    
                attempt += 1
                id_code = None
            
            if id_code is None:
                return Response({
                    'error': f'Unable to generate unique ID. Please try again.'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Create the student ID
            admin_id = AdminIssuedID.objects.create(
                id_code=id_code,
                role='student',
                assigned_name=assigned_name,
                assigned_phone=assigned_phone,
                notes=school_note,
                is_active=True,
                created_by=request.user  # School creates these
            )
            ids.append(AdminIssuedIDSerializer(admin_id).data)
        
        return Response({
            'ids': ids,
            'count': len(ids),
            'message': f'Successfully generated {len(ids)} student ID(s)'
        }, status=status.HTTP_201_CREATED)


# Volunteer: View assigned school participants
class VolunteerSchoolParticipantsView(generics.ListAPIView):
    serializer_class = SchoolParticipantSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.role != 'volunteer':
            return SchoolParticipant.objects.none()
        
        # Get schools assigned to this volunteer
        assignments = SchoolVolunteerAssignment.objects.filter(
            volunteer=self.request.user,
            is_active=True
        )
        school_ids = assignments.values_list('school_id', flat=True)
        
        return SchoolParticipant.objects.filter(school_id__in=school_ids)


# Volunteer: Verify student against school data
class VolunteerVerifyStudentView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    
    def create(self, request):
        if request.user.role != 'volunteer':
            raise PermissionDenied("Only volunteers can verify students")
        
        participant_id = request.data.get('participant_id')
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        
        # Find school participant
        assignments = SchoolVolunteerAssignment.objects.filter(
            volunteer=request.user,
            is_active=True
        )
        school_ids = assignments.values_list('school_id', flat=True)
        
        try:
            school_participant = SchoolParticipant.objects.filter(
                school_id__in=school_ids,
                participant_id=participant_id,
                first_name__iexact=first_name,
                last_name__iexact=last_name
            ).first()
            
            if not school_participant:
                return Response({'error': 'Participant not found in school data'}, 
                              status=status.HTTP_404_NOT_FOUND)
            
            # Mark as verified
            school_participant.verified_by_volunteer = True
            from django.utils import timezone
            school_participant.verified_at = timezone.now()
            school_participant.volunteer = request.user
            school_participant.save()
            
            return Response({'message': 'Participant verified', 
                            'participant': SchoolParticipantSerializer(school_participant).data})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# School: View own submitted participants
class SchoolViewOwnParticipantsView(generics.ListAPIView):
    serializer_class = SchoolParticipantSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.role != 'school':
            return SchoolParticipant.objects.none()
        
        # Return participants submitted by this school
        return SchoolParticipant.objects.filter(school=self.request.user).order_by('-submitted_at')


# Admin: Assign volunteers to schools
class AdminAssignVolunteerToSchoolView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    
    def create(self, request):
        school_id = request.data.get('school_id')
        volunteer_id = request.data.get('volunteer_id')
        
        try:
            school = User.objects.get(id=school_id, role='school')
            volunteer = User.objects.get(id=volunteer_id, role='volunteer')
        except User.DoesNotExist:
            return Response({'error': 'Invalid school or volunteer ID'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Create or update assignment
        assignment, created = SchoolVolunteerAssignment.objects.update_or_create(
            school=school,
            volunteer=volunteer,
            defaults={'assigned_by': request.user, 'is_active': True}
        )
        
        return Response({'message': 'Assignment created' if created else 'Assignment updated',
                        'assignment': SchoolVolunteerAssignmentSerializer(assignment).data},
                       status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)


# Public: View school standings
class SchoolStandingsView(generics.ListAPIView):
    serializer_class = SchoolStandingSerializer
    permission_classes = [AllowAny]
    queryset = SchoolStanding.objects.all()


# Admin: List all issued IDs with filters
class AdminIssuedIDListView(generics.ListAPIView):
    serializer_class = AdminIssuedIDSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]
    
    def get_queryset(self):
        queryset = AdminIssuedID.objects.all().order_by('-created_at')
        
        # Filter by role
        role = self.request.query_params.get('role')
        if role:
            queryset = queryset.filter(role=role)
        
        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter == 'available':
            queryset = queryset.filter(is_used=False, is_active=True)
        elif status_filter == 'used':
            queryset = queryset.filter(is_used=True, is_verified=False)
        elif status_filter == 'verified':
            queryset = queryset.filter(is_verified=True)
        elif status_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        # Search by ID code or name
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(id_code__icontains=search) | 
                Q(assigned_name__icontains=search) |
                Q(used_by__username__icontains=search) |
                Q(used_by__email__icontains=search)
            )
        
        return queryset


# Admin: Update or deactivate an issued ID
class AdminIssuedIDDetailView(generics.RetrieveUpdateAPIView):
    queryset = AdminIssuedID.objects.all()
    serializer_class = AdminIssuedIDSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]
    
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        
        # Allow updating assigned_name, assigned_phone, notes, and is_active
        if 'assigned_name' in request.data:
            instance.assigned_name = request.data['assigned_name']
        
        if 'assigned_phone' in request.data:
            instance.assigned_phone = request.data['assigned_phone']
        
        if 'notes' in request.data:
            instance.notes = request.data['notes']
        
        if 'is_active' in request.data:
            # Prevent deactivating already-used IDs unless explicitly forced
            if instance.is_used and not request.data.get('force', False):
                return Response({
                    'error': 'Cannot deactivate an ID that has been used. Use force=true to override.'
                }, status=status.HTTP_400_BAD_REQUEST)
            instance.is_active = request.data['is_active']
        
        instance.save()
        
        return Response(AdminIssuedIDSerializer(instance).data)


# Admin: Check ID validity (public for registration form)
@api_view(['POST'])
@permission_classes([AllowAny])
def check_id_validity(request):
    """
    Public endpoint to check if an ID code is valid and available.
    Used by registration form for real-time validation.
    """
    id_code = request.data.get('id_code', '').strip().upper()
    
    if not id_code:
        return Response({'valid': False, 'error': 'ID code is required'}, 
                       status=status.HTTP_400_BAD_REQUEST)
    
    try:
        issued_id = AdminIssuedID.objects.get(id_code=id_code)
    except AdminIssuedID.DoesNotExist:
        return Response({
            'valid': False,
            'error': 'Invalid ID code'
        })
    
    # Check various validity conditions
    if not issued_id.is_active:
        return Response({
            'valid': False,
            'error': 'This ID has been deactivated'
        })
    
    if issued_id.is_used:
        return Response({
            'valid': False,
            'error': 'This ID has already been used'
        })
    
    # ID is valid and available
    return Response({
        'valid': True,
        'role': issued_id.role,
        'assigned_name': issued_id.assigned_name,
        'message': f'Valid {issued_id.role} ID'
    })


# Admin: List all school participants (for user management panel)
class AdminSchoolParticipantsListView(generics.ListAPIView):
    """
    Admin endpoint to view all school participants pending approval.
    This displays in the User Management panel under Students section.
    """
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = SchoolParticipantSerializer

    def get_queryset(self):
        queryset = SchoolParticipant.objects.all().select_related('school').prefetch_related('events')

        # Filter by approval status (default: pending queue)
        status_filter = (self.request.query_params.get('status') or 'pending').strip().lower()
        if status_filter == 'pending':
            # Participants who don't have a user account yet
            queryset = queryset.filter(verified_by_volunteer=False)
        elif status_filter == 'approved':
            # Participants who have been approved
            queryset = queryset.filter(verified_by_volunteer=True)

        # Filter by section (LP/UP/HS/HSS)
        section = self.request.query_params.get('section', None)
        if section:
            if section == 'LP':
                queryset = queryset.filter(student_class__gte=1, student_class__lte=3)
            elif section == 'UP':
                queryset = queryset.filter(student_class__gte=4, student_class__lte=7)
            elif section == 'HS':
                queryset = queryset.filter(student_class__gte=8, student_class__lte=10)
            elif section == 'HSS':
                queryset = queryset.filter(student_class__gte=11, student_class__lte=12)

        # Filter by school
        school_id = self.request.query_params.get('school', None)
        if school_id:
            queryset = queryset.filter(school_id=school_id)

        return queryset.order_by('-submitted_at')


class AdminSchoolGroupParticipantsListView(generics.ListAPIView):
    """
    Admin endpoint to view submitted school group entries.
    """
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = SchoolGroupEntrySerializer

    def get_queryset(self):
        queryset = SchoolGroupEntry.objects.all().select_related(
            'school',
            'leader_user',
            'reviewed_by',
        ).prefetch_related(
            'members',
            'events',
        )

        status_filter_raw = str(
            _read_compat_field(self.request.query_params, 'status', 'statusFilter')
            or _read_compat_field(self.request.query_params, 'approval_status', 'approvalStatus')
            or ''
        ).strip().lower()
        status_aliases = {
            'pending': 'pending',
            'pend': 'pending',
            'approved': 'approved',
            'approve': 'approved',
            'accept': 'approved',
            'accepted': 'approved',
            'rejected': 'rejected',
            'reject': 'rejected',
            'declined': 'rejected',
            'all': 'all',
        }
        if not status_filter_raw:
            status_filter = 'pending'
        else:
            status_filter = status_aliases.get(status_filter_raw, status_filter_raw)
        if status_filter not in {'pending', 'approved', 'rejected', 'all'}:
            status_filter = 'pending'

        if status_filter in {'pending', 'approved', 'rejected'}:
            queryset = queryset.filter(status=status_filter)

        group_class = str(
            _read_compat_field(self.request.query_params, 'group_class', 'groupClass')
            or _read_compat_field(self.request.query_params, 'class', 'groupLevel')
            or ''
        ).strip().upper()
        if group_class in GROUP_CLASS_CHOICES:
            queryset = queryset.filter(group_class=group_class)

        gender_raw = str(
            _read_compat_field(self.request.query_params, 'gender_category', 'genderCategory')
            or _read_compat_field(self.request.query_params, 'gender', 'genderFilter')
            or ''
        ).strip().upper()
        gender_aliases = {
            'BOYS': 'BOYS',
            'GIRLS': 'GIRLS',
            'MIXED': 'MIXED',
            'MALE': 'BOYS',
            'FEMALE': 'GIRLS',
        }
        gender_category = gender_aliases.get(gender_raw, gender_raw)
        if gender_category in GROUP_GENDER_CHOICES:
            queryset = queryset.filter(gender_category=gender_category)

        school_id = _read_compat_field(self.request.query_params, 'school', 'schoolId')
        if school_id in [None, '']:
            school_id = _read_compat_field(self.request.query_params, 'school_id', 'schoolUserId')
        if school_id:
            queryset = queryset.filter(school_id=school_id)

        event_id = _read_compat_field(self.request.query_params, 'event', 'eventId')
        if event_id in [None, '']:
            event_id = self.request.query_params.get('event_id')
        if event_id:
            queryset = queryset.filter(events__id=event_id)

        search = str(
            _read_compat_field(self.request.query_params, 'search', 'q')
            or self.request.query_params.get('query')
            or ''
        ).strip()
        if search:
            queryset = queryset.filter(
                Q(group_id__icontains=search) |
                Q(leader_full_name__icontains=search) |
                Q(school__username__icontains=search) |
                Q(school__school__name__icontains=search)
            )

        return queryset.order_by('-submitted_at').distinct()


@extend_schema(
    request=AdminSchoolGroupApproveRequestSerializer,
    responses={
        200: AdminSchoolGroupApproveResponseSerializer,
        400: OpenApiResponse(response=ErrorEnvelopeSerializer),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
        409: OpenApiResponse(response=ErrorEnvelopeSerializer),
        500: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_approve_school_group_participant(request, group_entry_id):
    request_serializer = AdminSchoolGroupApproveRequestSerializer(data=request.data)
    request_serializer.is_valid(raise_exception=True)

    try:
        group_entry = SchoolGroupEntry.objects.select_related(
            'school',
            'leader_user',
        ).prefetch_related('members', 'events').get(pk=group_entry_id)
    except SchoolGroupEntry.DoesNotExist:
        return _error_response('Group entry not found', status.HTTP_404_NOT_FOUND, code='not_found')

    if group_entry.status == 'approved':
        return _error_response(
            'Group entry has already been approved',
            status.HTTP_400_BAD_REQUEST,
            code='already_approved',
            details={'group': SchoolGroupEntrySerializer(group_entry, context={'request': request}).data},
        )

    leader_member = group_entry.members.filter(is_leader=True).order_by('member_order').first()
    if leader_member is None:
        return _error_response(
            'Leader details are missing for this group entry',
            status.HTTP_400_BAD_REQUEST,
            code='missing_leader',
        )

    resolved_leader_user = group_entry.leader_user
    if resolved_leader_user is None:
        resolved_leader_user = _resolve_group_leader_user(
            group_entry.school,
            leader_member.first_name,
            leader_member.last_name,
        )

    try:
        leader_student_class = _default_student_class_for_group_class(group_entry.group_class)
    except ValueError as exc:
        return _error_response(str(exc), status.HTTP_400_BAD_REQUEST, code='invalid_group_class')

    group_registration_id = _school_group_leader_registration_id(group_entry.pk)
    legacy_group_registration_id = f"SG-{group_entry.school_id}-{group_entry.group_id}"
    group_gender = group_entry.gender_category if group_entry.gender_category in {'BOYS', 'GIRLS'} else None

    try:
        provisioned_account = _provision_school_student_account(
            school_user=group_entry.school,
            first_name=leader_member.first_name,
            last_name=leader_member.last_name,
            student_class=leader_student_class,
            gender=group_gender,
            username_seed=group_entry.group_id,
            primary_registration_id=group_registration_id,
            legacy_registration_ids=[legacy_group_registration_id, group_entry.group_id],
            existing_user_hint=resolved_leader_user,
            reassign_registration_id_values={None, '', legacy_group_registration_id, group_entry.group_id},
        )
    except ValueError as exc:
        return _error_response(str(exc), status.HTTP_400_BAD_REQUEST, code='invalid_state')
    except IntegrityError as exc:
        return _error_response(
            'Failed to create group leader account: username or other unique field already exists.',
            status.HTTP_409_CONFLICT,
            code='conflict',
            details={'db_error': str(exc)},
        )
    except Exception as exc:
        return _error_response(
            f'Failed to create group leader account: {str(exc)}',
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            code='server_error',
        )

    group_entry.leader_user = provisioned_account['user']
    group_entry.status = 'approved'
    group_entry.reviewed_at = timezone.now()
    group_entry.reviewed_by = request.user
    notes = str(request_serializer.validated_data.get('notes') or '').strip()
    if notes:
        group_entry.review_notes = notes
    group_entry.save(update_fields=['status', 'reviewed_at', 'reviewed_by', 'review_notes', 'leader_user', 'updated_at'])

    try:
        if group_entry.school.email:
            subject = f'Group Entry Approved: {group_entry.group_id}'
            message = f'''Dear {group_entry.school.username},

Your group entry has been approved and login credentials were generated for the group leader.

Group Details:
Group ID: {group_entry.group_id}
Group Class: {group_entry.group_class}
Leader: {leader_member.first_name} {leader_member.last_name}

Group Leader Login Credentials:
Username: {provisioned_account["username"]}
Password: {provisioned_account["password"]}

Please share these credentials only with the approved group leader.

Best regards,
E-Kalolsavam Admin Team
'''
            send_mail(subject, message, 'noreply@ekalolsavam.com', [group_entry.school.email], fail_silently=True)
    except Exception:
        pass

    return Response({
        'message': 'Group entry approved successfully',
        'group': SchoolGroupEntrySerializer(group_entry, context={'request': request}).data,
        'user_credentials': {
            'username': provisioned_account['username'],
            'password': provisioned_account['password'],
            'section': provisioned_account['section'],
            'user_id': provisioned_account['user'].id,
        }
    }, status=status.HTTP_200_OK)


@extend_schema(
    request=AdminSchoolGroupRejectRequestSerializer,
    responses={
        200: AdminSchoolGroupRejectResponseSerializer,
        400: OpenApiResponse(response=ErrorEnvelopeSerializer),
        403: OpenApiResponse(response=ErrorEnvelopeSerializer),
        404: OpenApiResponse(response=ErrorEnvelopeSerializer),
    },
)
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_reject_school_group_participant(request, group_entry_id):
    request_serializer = AdminSchoolGroupRejectRequestSerializer(data=request.data)
    request_serializer.is_valid(raise_exception=True)

    try:
        group_entry = SchoolGroupEntry.objects.select_related('school').prefetch_related('members', 'events').get(pk=group_entry_id)
    except SchoolGroupEntry.DoesNotExist:
        return _error_response('Group entry not found', status.HTTP_404_NOT_FOUND, code='not_found')

    if group_entry.status == 'rejected':
        return _error_response(
            'Group entry has already been rejected',
            status.HTTP_400_BAD_REQUEST,
            code='already_rejected',
            details={'group': SchoolGroupEntrySerializer(group_entry, context={'request': request}).data},
        )

    review_notes = str(
        request_serializer.validated_data.get('notes')
        or request_serializer.validated_data.get('reason')
        or ''
    ).strip()
    if not review_notes:
        return _error_response(
            'notes or reason is required to reject a group entry',
            status.HTTP_400_BAD_REQUEST,
            code='missing_review_notes',
        )

    group_entry.status = 'rejected'
    group_entry.review_notes = review_notes
    group_entry.reviewed_at = timezone.now()
    group_entry.reviewed_by = request.user
    group_entry.save(update_fields=['status', 'review_notes', 'reviewed_at', 'reviewed_by', 'updated_at'])

    return Response({
        'message': 'Group entry rejected successfully',
        'group': SchoolGroupEntrySerializer(group_entry, context={'request': request}).data,
    }, status=status.HTTP_200_OK)


# Admin: Approve school participant and create user account
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_approve_school_participant(request, participant_id):
    """
    Approve a school participant and create their user account.
    Generates username/password and assigns section based on class.
    """
    try:
        participant = SchoolParticipant.objects.select_related('school').get(pk=participant_id)
    except SchoolParticipant.DoesNotExist:
        return Response({'error': 'Participant not found'}, status=status.HTTP_404_NOT_FOUND)

    # Check if already approved
    if participant.verified_by_volunteer:
        return Response({
            'error': 'Participant has already been approved',
            'participant': SchoolParticipantSerializer(participant).data
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        participant_school = getattr(participant.school, 'school', None)
        if participant_school is None:
            participant_school = _ensure_school_user_linked_to_school_profile(participant.school)
        if participant_school is None:
            return Response({
                'error': 'School account is not linked to a School profile. Link the school user to a School before approving participants.',
                'school_user': {
                    'id': participant.school.id,
                    'username': participant.school.username
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        school_registration_id = _school_participant_registration_id(participant.pk)
        legacy_registration_id = f"SP-{participant_school.id}-{participant.participant_id}"

        with transaction.atomic():
            provisioned_account = _provision_school_student_account(
                school_user=participant.school,
                first_name=participant.first_name,
                last_name=participant.last_name,
                student_class=participant.student_class,
                gender=participant.gender,
                username_seed=participant.participant_id,
                primary_registration_id=school_registration_id,
                legacy_registration_ids=[legacy_registration_id, participant.participant_id],
                reassign_registration_id_values={None, '', participant.participant_id, legacy_registration_id},
            )
            user = provisioned_account['user']
            username = provisioned_account['username']
            password = provisioned_account['password']
            section = provisioned_account['section']

            # Mark participant as verified
            participant.verified_by_volunteer = True
            participant.verified_at = timezone.now()
            if getattr(request.user, 'role', None) == 'volunteer':
                participant.volunteer = request.user
            participant.save()

        # Send email notification to school (if school has email)
        try:
            if participant.school.email:
                subject = f'Participant Approved: {participant.first_name} {participant.last_name}'
                message = f'''Dear {participant.school.username},

Your participant has been approved and their account has been created.

Participant Details:
Name: {participant.first_name} {participant.last_name}
Participant ID: {participant.participant_id}
Class: {participant.student_class}
Section: {section}

Login Credentials:
Username: {username}
Password: {password}

The participant can now log in to the system and register for events.

Best regards,
E-Kalolsavam Admin Team
'''
                send_mail(subject, message, 'noreply@ekalolsavam.com', [participant.school.email], fail_silently=True)
        except Exception as e:
            # Don't fail the approval if email sending fails
            print(f"Failed to send email: {str(e)}")

        return Response({
            'message': 'Participant approved successfully',
            'participant': SchoolParticipantSerializer(participant).data,
            'user_credentials': {
                'username': username,
                'password': password,  # Only returned once
                'section': section,
                'user_id': user.id
            }
        }, status=status.HTTP_201_CREATED)

    except ValueError as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
    except IntegrityError as e:
        return Response({
            'error': 'Failed to create user account: username or other unique field already exists.',
            'details': str(e)
        }, status=status.HTTP_409_CONFLICT)
    except Exception as e:
        return Response({
            'error': f'Failed to create user account: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Admin: Bulk approve school participants
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def admin_bulk_approve_school_participants(request):
    """
    Bulk approve multiple school participants at once.
    """
    participant_ids = request.data.get('participant_ids', [])

    if not participant_ids or not isinstance(participant_ids, list):
        return Response({
            'error': 'participant_ids must be a non-empty list'
        }, status=status.HTTP_400_BAD_REQUEST)

    results = {
        'approved': [],
        'failed': [],
        'already_approved': []
    }

    for participant_id in participant_ids:
        try:
            participant = SchoolParticipant.objects.select_related('school').get(pk=participant_id)

            # Skip if already approved
            if participant.verified_by_volunteer:
                results['already_approved'].append({
                    'id': participant_id,
                    'participant_id': participant.participant_id,
                    'name': f"{participant.first_name} {participant.last_name}"
                })
                continue

            # Determine section
            student_class = participant.student_class
            if 1 <= student_class <= 3:
                section = 'LP'
            elif 4 <= student_class <= 7:
                section = 'UP'
            elif 8 <= student_class <= 10:
                section = 'HS'
            elif 11 <= student_class <= 12:
                section = 'HSS'
            else:
                results['failed'].append({
                    'id': participant_id,
                    'error': f'Invalid class: {student_class}'
                })
                continue

            password = secrets.token_urlsafe(12)[:12]
            temp_password_payload = signing.dumps({"p": password, "generated_at": timezone.now().isoformat()})

            participant_school = getattr(participant.school, 'school', None)
            if participant_school is None:
                participant_school = _ensure_school_user_linked_to_school_profile(participant.school)
            if participant_school is None:
                results['failed'].append({
                    'id': participant_id,
                    'error': 'School account is not linked to a School profile'
                })
                continue

            school_registration_id = _school_participant_registration_id(participant.pk)
            legacy_registration_id = f"SP-{participant_school.id}-{participant.participant_id}"

            existing_user = User.objects.filter(
                role='student',
                school=participant_school,
                registration_id=school_registration_id
            ).first()
            if existing_user is None:
                existing_user = User.objects.filter(
                    role='student',
                    school=participant_school,
                    registration_id=legacy_registration_id
                ).first()
            if existing_user is None:
                existing_user = User.objects.filter(
                    role='student',
                    school=participant_school,
                    registration_id=participant.participant_id
                ).first()
            if existing_user is None:
                existing_user = User.objects.filter(
                    role='student',
                    school=participant_school,
                    username=participant.participant_id.lower()
                ).first()

            if existing_user is not None:
                with transaction.atomic():
                    user = existing_user
                    user.set_password(password)
                    user.must_reset_password = True
                    user.temporary_password_encrypted = temp_password_payload
                    user.first_name = participant.first_name
                    user.last_name = participant.last_name
                    user.student_class = participant.student_class
                    user.gender = participant.gender
                    user.school_category_extra = section
                    user.is_active = True
                    user.approval_status = 'approved'
                    if user.registration_id in [None, '', participant.participant_id, legacy_registration_id]:
                        user.registration_id = school_registration_id
                    user.save()
                    username = user.username

                    # Mark as verified
                    participant.verified_by_volunteer = True
                    participant.verified_at = timezone.now()
                    if getattr(request.user, 'role', None) == 'volunteer':
                        participant.volunteer = request.user
                    participant.save()
            else:
                last_error = None
                for _ in range(5):
                    username = _generate_unique_student_username(participant.participant_id, participant_school.id)
                    try:
                        with transaction.atomic():
                            user = User.objects.create_user(
                                username=username,
                                password=password,
                                email=_student_placeholder_email(username),
                                first_name=participant.first_name,
                                last_name=participant.last_name,
                                role='student',
                                student_class=participant.student_class,
                                gender=participant.gender,
                                school_category_extra=section,  # Fixed: use school_category_extra instead of section
                                school=participant_school,
                                registration_id=school_registration_id,
                                is_active=True,
                                approval_status='approved',
                                must_reset_password=True,
                                temporary_password_encrypted=temp_password_payload,
                            )

                            # Mark as verified
                            participant.verified_by_volunteer = True
                            participant.verified_at = timezone.now()
                            if getattr(request.user, 'role', None) == 'volunteer':
                                participant.volunteer = request.user
                            participant.save()

                        last_error = None
                        break
                    except IntegrityError as e:
                        last_error = e
                        continue

                if last_error is not None:
                    raise last_error

            results['approved'].append({
                'id': participant_id,
                'participant_id': participant.participant_id,
                'name': f"{participant.first_name} {participant.last_name}",
                'username': username,
                'password': password,
                'section': section
            })

        except SchoolParticipant.DoesNotExist:
            results['failed'].append({
                'id': participant_id,
                'error': 'Participant not found'
            })
        except IntegrityError as e:
            results['failed'].append({
                'id': participant_id,
                'error': 'Username or other unique field already exists',
                'details': str(e)
            })
        except Exception as e:
            results['failed'].append({
                'id': participant_id,
                'error': str(e)
            })

    return Response({
        'message': f'Approved {len(results["approved"])} participants',
        'results': results
    }, status=status.HTTP_200_OK)
